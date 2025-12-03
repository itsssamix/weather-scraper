import time #biblioteca time para manipulação de data e hora p/ registrar momento da coleta
from openpyxl import Workbook, load_workbook #biblioteca para manipular excel

def salvar_temperatura(temperatura, arquivo="historico_temperatura.xlsx"):
    """Salva data/hora e temperatura no Excel."""

    try:
        wb = load_workbook(arquivo)
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active
        ws.append(["Data/Hora", "Temperatura"])

    data_hora = time.strftime("%Y-%m-%d %H:%M:%S") #formata data como string
    ws.append([data_hora, temperatura]) #adiciona nova linha na planilha

    wb.save(arquivo) #salva arquivo